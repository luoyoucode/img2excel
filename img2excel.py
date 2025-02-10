import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import threading
import queue


class ColorQuantizer:
    @staticmethod
    def quantize_color(rgb, bits=5):
        """颜色量化方法（将颜色值压缩到指定位数）"""
        shift = 8 - bits
        return tuple((c >> shift) << shift for c in rgb)


class ExcelSizeValidator:
    MAX_STYLES = 64000  # Excel样式数量安全阈值
    MAX_DIMENSIONS = (1048576, 16384)  # 最大行数和列数

    @classmethod
    def validate_size(cls, width, height):
        """验证图片尺寸是否超出Excel限制"""
        max_rows, max_cols = cls.MAX_DIMENSIONS
        if height > max_rows:
            return False, f"图片高度({height})超过Excel最大行数限制({max_rows})"
        if width > max_cols:
            return False, f"图片宽度({width})超过Excel最大列数限制({max_cols})"
        return True, ""


class ImageToExcelConverter:
    def __init__(self, master):
        self.master = master
        master.title("图片转Excel专业版")
        master.geometry("600x400")

        # 初始化多线程相关组件
        self.queue = queue.Queue()
        self.is_running = False
        self.quantization_bits = 5  # 默认颜色量化位数

        # 获取桌面路径
        self.desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

        # 创建GUI组件
        self.create_widgets()
        self.master.after(100, self.process_queue)

    def create_widgets(self):
        """创建所有界面组件"""
        # 图片路径组件
        tk.Label(self.master, text="图片路径:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.image_entry = tk.Entry(self.master, width=50)
        self.image_entry.insert(0, self.desktop_path)
        self.image_entry.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.master, text="浏览", command=self.browse_image).grid(row=0, column=2, padx=5)

        # 输出路径组件
        tk.Label(self.master, text="输出文件夹:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.output_entry = tk.Entry(self.master, width=50)
        self.output_entry.insert(0, self.desktop_path)
        self.output_entry.grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.master, text="浏览", command=self.browse_output).grid(row=1, column=2, padx=5)

        # 颜色量化设置
        tk.Label(self.master, text="颜色精度:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.bits_scale = tk.Scale(self.master, from_=1, to=8, orient=tk.HORIZONTAL,
                                   command=self.update_quantization_bits)
        self.bits_scale.set(self.quantization_bits)
        self.bits_scale.grid(row=2, column=1, columnspan=2, sticky="ew", padx=5)
        tk.Label(self.master, text="低 ◄───────► 高").grid(row=3, column=1, sticky="w")

        # 进度条
        self.progress = ttk.Progressbar(self.master, length=500, mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=3, pady=10, padx=5)

        # 状态标签
        self.status_label = tk.Label(self.master, text="准备就绪", anchor="w")
        self.status_label.grid(row=5, column=0, columnspan=3, sticky="ew", padx=5)

        # 转换按钮
        self.convert_btn = tk.Button(self.master, text="开始转换", command=self.start_conversion,
                                     bg="#4CAF50", fg="white", state=tk.NORMAL)
        self.convert_btn.grid(row=6, column=1, pady=10)

    def browse_image(self):
        """选择图片文件"""
        file_path = filedialog.askopenfilename(
            initialdir=self.desktop_path,
            filetypes=[("图片文件", "*.png;*.jpg;*.jpeg;*.bmp")]
        )
        if file_path:
            self.image_entry.delete(0, tk.END)
            self.image_entry.insert(0, file_path)

    def browse_output(self):
        """选择输出目录"""
        dir_path = filedialog.askdirectory(initialdir=self.desktop_path)
        if dir_path:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, dir_path)

    def update_quantization_bits(self, value):
        """更新颜色量化位数"""
        self.quantization_bits = int(value)

    def start_conversion(self):
        """开始转换"""
        if self.is_running:
            return

        self.is_running = True
        self.convert_btn.config(state=tk.DISABLED, text="转换中...")
        self.status_label.config(text="正在初始化...")
        self.progress["value"] = 0

        conversion_thread = threading.Thread(target=self.conversion_worker, daemon=True)
        conversion_thread.start()

    def conversion_worker(self):
        """实际执行转换的工作线程"""
        try:
            image_path = self.image_entry.get()
            output_dir = self.output_entry.get()

            # 验证输入参数
            if not all([image_path, output_dir]):
                self.queue.put(("error", "请填写所有路径"))
                return

            if not os.path.exists(image_path):
                self.queue.put(("error", "图片文件不存在"))
                return

            # 打开并处理图片
            with Image.open(image_path) as img:
                img = img.convert("RGB")
                width, height = img.size

                # 验证尺寸限制
                is_valid, msg = ExcelSizeValidator.validate_size(width, height)
                if not is_valid:
                    self.queue.put(("error", msg))
                    return

                # 创建Excel工作簿
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "Pixel Art"

                # 设置单元格尺寸
                for col in range(1, width + 1):
                    sheet.column_dimensions[get_column_letter(col)].width = 2.75
                for row in range(1, height + 1):
                    sheet.row_dimensions[row].height = 15

                # 初始化颜色缓存
                color_cache = {}
                style_count = 0
                warning_shown = False
                pixels = img.load()
                total_pixels = width * height
                processed = 0

                # 处理每个像素
                for y in range(height):
                    for x in range(width):
                        # 获取并量化颜色
                        r, g, b = pixels[x, y]
                        quantized_rgb = ColorQuantizer.quantize_color((r, g, b), self.quantization_bits)
                        hex_color = f"FF{quantized_rgb[0]:02X}{quantized_rgb[1]:02X}{quantized_rgb[2]:02X}"

                        # 管理样式缓存
                        if hex_color not in color_cache:
                            if style_count >= ExcelSizeValidator.MAX_STYLES:
                                if not warning_shown:
                                    self.queue.put(("warning", "达到样式数量限制，开始合并相似颜色"))
                                    warning_shown = True
                                # 查找最近似颜色
                                hex_color = min(color_cache.keys(),
                                                key=lambda k: self.color_distance(k, hex_color))
                            else:
                                color_cache[hex_color] = PatternFill(
                                    start_color=hex_color,
                                    end_color=hex_color,
                                    fill_type="solid"
                                )
                                style_count += 1

                        # 应用单元格颜色
                        sheet.cell(row=y + 1, column=x + 1).fill = color_cache[hex_color]
                        processed += 1

                        # 更新进度（每处理0.5%更新一次）
                        if processed % (total_pixels // 200) == 0:
                            self.queue.put(("progress", processed / total_pixels * 100))

                # 保存文件
                filename = f"{os.path.splitext(os.path.basename(image_path))[0]}_pixel.xlsx"
                output_path = os.path.join(output_dir, filename)
                wb.save(output_path)

                # 发送完成消息
                self.queue.put(("success", (output_path, style_count)))

        except Exception as e:
            self.queue.put(("error", str(e)))

    def color_distance(self, hex1, hex2):
        """计算两个颜色之间的差异"""

        def hex_to_rgb(h):
            return tuple(int(h[i:i + 2], 16) for i in (2, 4, 6))

        rgb1 = hex_to_rgb(hex1)
        rgb2 = hex_to_rgb(hex2)
        return sum((a - b) ** 2 for a, b in zip(rgb1, rgb2))

    def process_queue(self):
        """处理消息队列"""
        try:
            while True:
                msg_type, content = self.queue.get_nowait()

                if msg_type == "progress":
                    self.progress["value"] = content
                    self.status_label.config(text=f"处理进度：{content:.1f}%")
                elif msg_type == "success":
                    output_path, style_count = content
                    messagebox.showinfo("转换成功",
                                        f"文件已保存到：{output_path}\n"
                                        f"使用样式数：{style_count}/{ExcelSizeValidator.MAX_STYLES}")
                    self.reset_ui()
                elif msg_type == "error":
                    messagebox.showerror("转换失败", content)
                    self.reset_ui()
                elif msg_type == "warning":
                    messagebox.showwarning("颜色合并警告", content)

        except queue.Empty:
            pass
        finally:
            self.master.after(100, self.process_queue)

    def reset_ui(self):
        """重置界面状态"""
        self.is_running = False
        self.convert_btn.config(state=tk.NORMAL, text="开始转换")
        self.progress["value"] = 0
        self.status_label.config(text="准备就绪")


if __name__ == "__main__":
    root = tk.Tk()
    app = ImageToExcelConverter(root)
    root.mainloop()